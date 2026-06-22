import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Mappings for brand normalization
BRAND_MAPPING = {
    "ariel": "Ariel",
    "bold": "Bold",
    "daz": "Daz",
    "eco-egg": "Ecoegg",
    "elbow_grease": "Elbow Grease",
    "fairy": "Fairy",
    "the_pink_stuff": "The Pink Stuff",
    "vanish": "Vanish",
    "g.bellini": "G. Bellini",
    "cif": "Cif",
    "crystale": "Crystale",
    "domestos": "Domestos",
    "dr.beckmann": "Dr. Beckmann",
    "fabulosa": "Fabulosa",
    "kilrock": "Kilrock",
    "limescale_preventer": "Limescale Preventer",
    "max_flush": "Max Flush",
    "mighty_power": "Mighty Power",
    "power_force": "Power Force",
    "w5": "W5",
    "almat": "Almat",
    "dylon": "Dylon",
    "laundry_fresh": "Laundry Fresh",
    "lenor": "Lenor",
    "paclan": "Paclan",
    "sparkle_and_shine": "Sparkle and Shine",
    "maybelline_new_york": "Maybelline New York",
    "american_touch": "American Touch",
    "cien": "Cien",
    "cussons": "Cussons",
    "dentalux": "Dentalux",
    "imperial_leather": "Imperial Leather",
    "lacura": "Lacura",
    "pears": "Pears",
    "purple": "Purple",
    "sensodyne": "Sensodyne",
    "true_smile": "True Smile",
    "wisdom": "Wisdom",
    "beautifully_scrumptious": "Beautifully Scrumptious",
    "carmex": "Carmex",
    "white_whitening": "White Whitening",
    "dermav10": "Derma V10",
    "loreal_paris": "L'Oréal Paris",
    "sun_oasis": "Sun Oasis",
    "tee_tree": "Tea Tree",
    "activ_max": "Activ Max",
    "allok": "Allok",
    "go_nutrition": "Go Nutrition",
    "haliborange": "Haliborange",
    "healthpoint": "Healthpoint",
    "kirkland_signature": "Kirkland Signature",
    "major": "Major",
    "minavit": "MinaVit",
    "myvitamins": "Myvitamins",
    "nutrilife": "NutriLife",
    "vitaminstore": "VitaminStore",
    "wellwoman": "Wellwoman"
}

# Known variants list to help with extraction
KNOWN_VARIANTS = [
    "original", "lavender_and_camomile", "lavender", "pink_blossom", "cherry_blossom",
    "lime", "magnolia", "pink_magnolia", "lemon", "citrus_sparkle", "citrus", "floral",
    "ocean", "blue_water_and_cleanliness", "blue_water", "fresh_citrus", "fresh_scent",
    "blossom", "coconut", "vibrant", "spring_awakening", "summer_breeze", "northern_solstice",
    "seabreeze_stroll", "105explorer", "20pioneer", "320individualist", "325shot_caller",
    "90huntress", "105_explorer", "20_pioneer", "320_individualist", "325_shot_caller",
    "90_huntress", "extreme_dry", "comfort_fresh", "intensive", "invisible", "exotic_fragrance",
    "vitalizing_fragrance", "moisture_plus", "shea_butter_and_honey", "space_candy_and_coconut_cream",
    "oatmeal", "sea-salt_and_pomegranate", "noir", "sensitive", "aloe_vera", "mint", "hemp_oil",
    "passion_fruit", "strawberry", "vanilla", "cherry", "classic", "caviar_illumination",
    "collagen_and_gold", "q10_renew", "rejuvelate", "kids", "junior", "deep", "homme_sport",
    "one_fragrance", "orange_flavour", "blackcurrant_flavour", "energy+", "mobility+"
]

# SKU prefixes for each category
SKU_PREFIXES = {
    "Detergents": "BC-DET-",
    "Fragrances": "BC-FRG-",
    "Household": "BC-HSH-",
    "Laundry_Care": "BC-LDR-",
    "Make-up": "BC-MKP-",
    "Personal_Care": "BC-PRC-",
    "Skin_Care": "BC-SKN-",
    "Supplements": "BC-SPL-"
}

# Regex to capture size patterns like: _25washes, _500ml, _100g, _4.25g, _30gummies, etc.
SIZE_REGEX = re.compile(
    r'_(\d+(?:\.\d+)?)(washes|Washes|g|ml|pack|sheets|tablets|wipes|capsules|chewable_tablets|soft_gels|gummies|pads|caplets|sachets)\b',
    re.IGNORECASE
)

# Walk backward to remove suffix ignoring underscores
def remove_suffix_ignore_underscores(text, suffix_clean):
    chars_to_remove = len(suffix_clean)
    idx = len(text) - 1
    removed_chars = 0
    while idx >= 0 and removed_chars < chars_to_remove:
        if text[idx] != '_':
            removed_chars += 1
        idx -= 1
    return text[:idx + 1].strip('_')

# Capitalization helper to format names professionally
def clean_capitalization(text):
    if not text:
        return ""
    words = text.split()
    cleaned = []
    
    # List of words that should remain lowercase unless at the beginning
    lowercase_words = {"in", "and", "or", "with", "for", "of", "to", "at", "by", "a", "an", "the"}
    
    # List of words that should be fully uppercase
    uppercase_words = {"q10", "xxl", "spf15", "spf20", "spf30", "spf50", "spf50+", "usp", "b6", "a+d", "sku"}
    
    # Predefined capitalization dictionary for mixed-case words
    special_cases = {
        "vitaminc": "Vitamin C",
        "vitamind": "Vitamin D",
        "vitamina": "Vitamin A",
        "omega3": "Omega-3",
        "omega-3": "Omega-3",
        "haliborange": "Haliborange",
        "minavit": "MinaVit",
        "myvitamins": "Myvitamins",
        "vitaminstore": "VitaminStore",
        "wellwoman": "Wellwoman"
    }
    
    for i, w in enumerate(words):
        w_lower = w.lower()
        if w_lower in lowercase_words:
            if i == 0:
                cleaned.append(w.capitalize())
            else:
                cleaned.append(w_lower)
        elif w_lower in uppercase_words or re.match(r'^spf\d+\+?$', w_lower):
            cleaned.append(w.upper())
        elif w_lower in special_cases:
            cleaned.append(special_cases[w_lower])
        elif w_lower.startswith("vitamin") and len(w_lower) > 7:
            letter = w_lower[7:].upper()
            cleaned.append(f"Vitamin {letter}")
        else:
            # Check for pattern like "VitaminC" -> "Vitamin C", "VitaminD" -> "Vitamin D"
            match = re.match(r'^(vitamin)([a-z])$', w_lower)
            if match:
                cleaned.append(f"Vitamin {match.group(2).upper()}")
            else:
                cleaned.append(w.capitalize())
            
    return ' '.join(cleaned)

# Helper function to generate descriptions based on components
def generate_product_description(brand, name, variant, category, qty, unit):
    brand_lower = brand.lower()
    name_lower = name.lower()
    variant_lower = variant.lower() if variant else ""
    s_phrase = f" ({qty} {unit})" if qty and unit else ""
    
    # 1. Custom matches for specific products
    if "carmex" in brand_lower:
        return f"Soothe and protect dry, chapped lips with Carmex Moisturising Lip Balm in {variant} scent{s_phrase}. Formulated with a unique cooling formula containing camphor and menthol for quick relief."
        
    if "maybelline" in brand_lower:
        return f"Super Stay Matte Ink liquid lipstick by Maybelline New York in shade {variant}. Offers a highly-pigmented, flawless matte finish that lasts up to 16 hours with zero transfer."
        
    if "ariel" in brand_lower:
        return f"Ariel's advanced laundry pods{s_phrase}. Designed to lift tough stains, brighten fabrics, and leave your clothes smelling exceptionally clean and fresh in one simple wash."
        
    if "bold" in brand_lower:
        if "lavender" in variant_lower or "lavender" in name_lower:
            return f"Bold All-in-One pods{s_phrase} infused with calming Lavender & Camomile. Combines deep cleaning power with built-in fabric softeners for cozy, fragrant laundry."
        if "pink blossom" in variant_lower or "pink" in variant_lower:
            return f"Bold All-in-One laundry pods{s_phrase} in sweet Pink Blossom scent. Delivers brilliant cleaning performance and long-lasting floral freshness while keeping fabrics soft."
        return f"Bold All-in-One laundry pods{s_phrase}. Features built-in fabric softeners and a premium scent profile to clean, soften, and freshen clothes in one step."
        
    if "daz" in brand_lower:
        if "cherry" in variant_lower or "cherry" in name_lower:
            return f"Daz XXL pods{s_phrase} in sweet Cherry Blossom. Provides a deep, brilliant clean for whites and colors, keeping your laundry bright and infused with a rich floral fragrance."
        return f"Daz All-in-One laundry pods{s_phrase}. Provides a reliable, brilliant clean for both whites and colored fabrics, leaving laundry smelling clean and fresh."
        
    if "eco-egg" in brand_lower or "ecoegg" in brand_lower:
        if "refill" in name_lower:
            return f"Ecoegg Laundry Egg refill pellets{s_phrase}. Gentle formula that is kind to skin, renewing your egg's washing power."
        return f"Ecoegg Laundry Egg{s_phrase}. Gentle formula and kind to skin for fresh, clean laundry."
        
    if "elbow grease" in brand_lower:
        return f"A powerful, targeted stain remover bar from Elbow Grease{s_phrase}. Specifically designed to pre-treat tough grease, oil, and food stains on clothing collars, cuffs, and household fabrics."
        
    if "fairy" in brand_lower:
        return f"Fairy Non-Biological laundry pods{s_phrase}. Dermatologically tested and approved for sensitive skin, offering gentle care, cleanliness, and huggable softness for baby clothing."
        
    if "the pink stuff" in brand_lower:
        if "oxi" in name_lower:
            return f"The Pink Stuff Miracle Oxi Stain Remover{s_phrase}. Uses active oxygen action to lift tough stains from fabrics and whites. Ideal for pre-treatments or adding directly to the wash."
        if "paste" in name_lower:
            return f"The Pink Stuff Miracle Cleaning Paste{s_phrase}, a tough household cleaner that is gentle on surfaces but tough on stains. Ideal for cleaning saucepans, cooktops, sinks, and tiles."
        if "cream" in name_lower:
            return f"The Pink Stuff Miracle Cream Cleaner{s_phrase}, a versatile liquid cream cleaner for hard surfaces. Easily penetrates and removes grease and grime, leaving a brilliant streak-free shine."
        if "toilet" in name_lower:
            return f"The Pink Stuff Toilet Cleaner Gel{s_phrase}. A thick, foaming cleaning gel that clings to toilet bowls to remove limescale and tough stains, leaving a sparkling clean finish."
            
    if "vanish" in brand_lower:
        if "crystal white" in name_lower or "white" in variant_lower:
            return f"Vanish Oxi Action Crystal White stain remover{s_phrase}. Specifically formulated to lift tough stains and restore whites to their brightest shade, even in cold washes."
        if "vibrant" in name_lower or "colour" in variant_lower or "colors" in variant_lower:
            return f"Vanish Oxi Action stain remover for vibrant colors{s_phrase}. A color-safe formula that removes stains on the first wash while protecting fabric colors from fading."
        if "bar" in name_lower:
            return f"Vanish Stain Remover Bar{s_phrase}. A targeted pre-treat stick that directly tackles stubborn spots like grease, wine, grass, and makeup on clothing and linens."
            
    if "cien" in brand_lower:
        if "q10" in name_lower:
            if "day" in name_lower:
                return f"Cien Q10 Intense Day Cream with SPF15{s_phrase}. Formulated with Coenzyme Q10 and Hyaluronic Acid to visibly reduce fine lines and protect skin from premature aging."
            if "night" in name_lower:
                return f"Cien Q10 Intense Night Cream{s_phrase}. Works overnight to restore and regenerate skin cells, deeply moisturizing and firming wrinkles with Q10 and Vitamin E."
            if "serum" in name_lower:
                return f"Cien Q10 Intense Hyaluronic Acid Serum{s_phrase}. A concentrated facial serum designed to deeply hydrate, plump fine lines, and boost skin elasticity for a youthful glow."
        if "facial wash" in name_lower or "face wash" in name_lower:
            if "moisturising" in name_lower or "moisture" in name_lower:
                return f"Cien Moisturising Facial Wash{s_phrase}. Gently cleanses and purifies skin, removing makeup and impurities while locking in moisture for a soft feel."
            return f"Cien Refreshing Facial Wash{s_phrase}. A light, soap-free gel cleanser that washes away dirt and excess oil, leaving skin feeling clean, fresh, and revitalized."
        if "scrub" in name_lower:
            return f"Cien Smoothing Facial Scrub{s_phrase}. Contains gentle exfoliating micro-particles to sweep away dead skin cells, refine skin texture, and reveal a radiant complexion."
        if "sun" in name_lower:
            return f"Cien Sun Suncare Lotion with high SPF50+ protection{s_phrase}. Offers broad-spectrum UVA/UVB protection, water-resistant formula, and deep hydration for sun-exposed skin."
        if "handcream" in name_lower or "hand cream" in name_lower:
            if "anti-ageing" in name_lower:
                return f"Cien Anti-Ageing Hand Cream{s_phrase}. Helps prevent signs of aging on hands, protecting skin from UV rays with coenzyme Q10 and maintaining skin elasticity."
            if "intensive" in name_lower:
                return f"Cien Intensive Hand Cream{s_phrase}. Provides quick relief and intense moisture for dry, rough, or chapped hands, restoring softness."
            return f"Cien Moisturising Hand Cream{s_phrase}. A fast-absorbing daily cream that keeps hands soft, hydrated, and protected from daily environmental stress."
        if "roll on" in name_lower or "deodorant" in name_lower:
            return f"Cien Deodorant Roll-On in {variant} scent{s_phrase}. Provides 48-hour odor protection, alcohol-free formula, and gentle care for sensitive underarm skin."
            
    if "lacura" in brand_lower:
        if "caviar" in name_lower:
            return f"Lacura Caviar Illumination Anti-Age Day Cream with SPF15{s_phrase}. A luxurious day cream containing caviar extract to stimulate cell renewal, reduce wrinkles, and firm skin."
        if "collagen" in name_lower:
            return f"Lacura Collagen and Gold Facial Serum{s_phrase}. A luxurious anti-aging serum that combines collagen and gold flakes to restore skin elasticity, plump lines, and add radiance."
        if "q10" in name_lower:
            return f"Lacura Q10 Renew Day Cream with SPF20{s_phrase}. Formulated with a Q10 peptide complex to protect skin from environmental damage and reduce the appearance of wrinkles."
        if "rejuvelate" in name_lower:
            return f"Lacura Rejuvelate Day Cream with SPF30{s_phrase}. Formulated to firm skin, combat wrinkles, and protect against UV damage, leaving the skin feeling hydrated and looking radiant."
        if "face wash" in name_lower:
            return f"Lacura Moisturising Face Wash{s_phrase}. Soap-free formula that cleanses skin without drying it out, leaving it soft and refreshed."
        if "wipes" in name_lower:
            return f"Lacura Facial Wipes{s_phrase}. Gently removes makeup, dirt, and waterproof mascara, leaving skin clean and refreshed. Suitable for daily use."
            
    if "minavit" in brand_lower:
        if "omega" in name_lower or "fish oil" in name_lower:
            return f"MinaVit Heart Health Omega-3 Fish Oil capsules{s_phrase}. Rich in EPA and DHA to support healthy cardiovascular function, normal brain performance, and visual health."
        if "vitamin c" in name_lower:
            return f"MinaVit Chewable Vitamin C tablets{s_phrase} in refreshing orange flavour. Supports a healthy immune system, reduces fatigue, and acts as a strong antioxidant."
        if "vitamin d" in name_lower:
            if "gummies" in name_lower:
                return f"MinaVit High Strength Vitamin D Gummies{s_phrase} in blackcurrant flavour. Support bone health, teeth, muscle function, and a strong immune system."
            return f"MinaVit Mobility and Bones High Strength Vitamin D capsules{s_phrase}. Supports the absorption of calcium, maintaining healthy bones and teeth."
        if "magnesium" in name_lower:
            return f"MinaVit Energy Magnesium with Vitamin B6 tablets{s_phrase}. Helps reduce tiredness, supports muscle function, and aids in maintaining a healthy nervous system."
        if "calcium" in name_lower:
            return f"MinaVit Mobility and Bones Calcium & Vitamin D tablets{s_phrase}. Formulated to help maintain strong bones, teeth, and support normal muscle function."
            
    if "vitaminstore" in brand_lower:
        if "cod liver oil" in name_lower:
            if "multivitamin" in name_lower:
                return f"VitaminStore High Strength Cod Liver Oil, Fish Oil, and Multivitamin capsules{s_phrase}. Supports joint mobility, cardiovascular health, and daily energy levels."
            return f"VitaminStore High Strength Cod Liver Oil and Fish Oil with Vitamin A & D{s_phrase}. Promotes healthy joint mobility, vision, and immune system function."
        if "evening primrose" in name_lower:
            return f"VitaminStore High Strength Evening Primrose Oil capsules{s_phrase}. Provides essential GLA fatty acids to support hormonal balance and maintain healthy skin."
        if "magnesium" in name_lower:
            return f"VitaminStore Magnesium & Vitamin B6 Energy+ tablets{s_phrase}. Supports energy-yielding metabolism, muscle recovery, and helps reduce daily fatigue."
            
    if "wellwoman" in brand_lower:
        return f"Wellwoman Original Vitabiotics capsules{s_phrase}. A comprehensive daily multivitamin formulated specifically for women's nutritional needs, containing B-vitamins, Iron, and Zinc."

    if "kirkland" in brand_lower:
        if "calcium" in name_lower:
            return f"Kirkland Signature Calcium, Magnesium & Zinc with Vitamin D{s_phrase}. Formulated to support strong bone structure, teeth, muscle contraction, and immune health."
        if "c 1000mg" in name_lower or "vitamin c" in name_lower:
            return f"Kirkland Signature Vitamin C 1000mg caplets{s_phrase}. Helps support a healthy immune system and provides high-potency antioxidant protection against free radicals."
            
    # Default fallbacks per category if no custom match
    v_str = f" in {variant}" if variant and variant.lower() not in ["original", "standard"] else ""
    if category == "Detergents":
        return f"Premium {brand} {name}{v_str}{s_phrase} laundry detergent. Expertly formulated to penetrate deep into fabric fibers, removing tough stains and keeping your clothes clean and fresh."
    elif category == "Fragrances":
        return f"Luxurious {brand} {name}{v_str}{s_phrase} fragrance. Features a long-lasting, sophisticated scent profile, perfect for daily wear or formal events."
    elif category == "Household":
        return f"Hygienic {brand} {name}{v_str}{s_phrase} cleaning solution. Easily dissolves tough grease, dirt, and stains around your home, leaving surfaces clean and sparkling."
    elif category == "Laundry_Care":
        return f"Premium fabric conditioner or enhancer: {brand} {name}{v_str}{s_phrase}. Keeps garments smelling incredible, static-free, and feeling exceptionally soft."
    elif category == "Make-up":
        v_phrase = f" in shade {variant}" if variant else ""
        return f"High-quality {brand} {name}{v_phrase}. Perfect for achieving a professional finish, this makeup product offers excellent pigmentation and smooth application."
    elif category == "Personal_Care":
        return f"Nourishing personal hygiene essential: {brand} {name}{v_str}{s_phrase}. Dermatologically tested to clean, protect, and refresh your skin daily."
    elif category == "Skin_Care":
        return f"Revitalizing skin care formula: {brand} {name}{v_str}{s_phrase}. Restores hydration, protects the skin barrier, and promotes a healthy, radiant complexion."
    elif category == "Supplements":
        return f"Daily health supplement {brand} {name}{v_str}{s_phrase}. Packed with vitamins and minerals to support your body's immune system, joint health, and overall well-being."
    else:
        return f"Authentic {brand} {name}{v_str}{s_phrase} product. Imported from the UK, offering premium quality and reliable performance for your daily needs."

def parse_filename(filename, category):
    # Strip extension
    base_name = os.path.splitext(filename)[0]
    
    # 1. Extract size / quantity
    qty, unit = None, None
    size_match = SIZE_REGEX.search(base_name)
    if size_match:
        qty = size_match.group(1)
        unit = size_match.group(2).lower()
        # Remove size match from base_name
        base_name = base_name[:size_match.start()] + base_name[size_match.end():]
    
    # Clean underscores and double spaces
    base_name = re.sub(r'_+', '_', base_name).strip('_')
    
    # 2. Extract Brand
    brand = None
    lower_base = base_name.lower()
    
    # Check brands sorted by length descending to match longest first
    sorted_brands = sorted(BRAND_MAPPING.keys(), key=len, reverse=True)
    for b_key in sorted_brands:
        if lower_base.startswith(b_key):
            brand = BRAND_MAPPING[b_key]
            # Remove brand from base_name
            base_name = base_name[len(b_key):].strip('_')
            break
            
    if not brand:
        # Fallback: take the first word as brand name
        parts = base_name.split('_')
        if parts:
            brand = parts[0].replace('-', ' ').title()
            base_name = '_'.join(parts[1:])
    
    # 3. Extract Variant
    variant = None
    lower_base = base_name.lower()
    
    # Check SPF ratings
    spf_match = re.search(r'spf\d+\+?', lower_base)
    if spf_match:
        variant = spf_match.group(0).upper()
        # Remove variant from base_name
        base_name = base_name[:spf_match.start()] + base_name[spf_match.end():]
    else:
        # Check known variants
        variant_matched = False
        sorted_variants = sorted(KNOWN_VARIANTS, key=len, reverse=True)
        for v_key in sorted_variants:
            v_key_clean = v_key.replace('_', '')
            base_name_clean = lower_base.replace('_', '')
            if base_name_clean.endswith(v_key_clean):
                variant = v_key.replace('_', ' ').title()
                # Remove suffix from base_name
                base_name = remove_suffix_ignore_underscores(base_name, v_key_clean)
                variant_matched = True
                break
                
    if not variant:
        # Default fallback variants based on common cases
        if "original" in lower_base:
            variant = "Original"
            base_name = base_name.replace("Original", "").replace("original", "").strip('_')
        elif "sensitive" in lower_base:
            variant = "Sensitive"
            base_name = base_name.replace("Sensitive", "").replace("sensitive", "").strip('_')
        else:
            variant = "Original"
            
    # Clean up product name
    product_name = base_name.replace('_', ' ').strip()
    product_name = re.sub(r'\s+', ' ', product_name)
    product_name = clean_capitalization(product_name)
    
    # Clean up variant
    if variant:
        variant = variant.replace('_', ' ')
        variant = clean_capitalization(variant)
        
    # Standardize Units for display
    unit_display = unit
    if unit:
        if unit == "g":
            unit_display = "g"
        elif unit == "ml":
            unit_display = "ml"
        elif unit == "washes":
            unit_display = "washes"
        elif unit == "tablets":
            unit_display = "tablets"
        elif unit == "gummies":
            unit_display = "gummies"
        elif unit == "wipes":
            unit_display = "wipes"
        elif unit == "sheets":
            unit_display = "sheets"
        elif unit == "capsules":
            unit_display = "capsules"
        elif unit == "chewable_tablets":
            unit_display = "chewable tablets"
        elif unit == "soft_gels":
            unit_display = "soft gels"
            
    return brand, product_name, variant, qty, unit_display

def main():
    # Detect which image folder name exists
    image_dir = "Product_Images_1"
    if not os.path.exists(image_dir) and os.path.exists("Product_Images"):
        image_dir = "Product_Images"
        
    excel_dir = "Inventory"
    excel_filename = os.path.join(excel_dir, "BritChoice_Product_Inventory.xlsx")
    
    if not os.path.exists(image_dir):
        print(f"Error: Neither '{image_dir}' nor 'Product_Images' directories exist.")
        return
        
    # Ensure the Excel output directory exists
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
        
    # 1. Read existing stock, prices & descriptions if the file already exists
    existing_stock_price = {}  # SKU -> (stock_val, price_val)
    existing_descriptions = {} # SKU -> description
    
    # Check for existing file in either the new path or the fallback root path
    check_paths = [excel_filename, "BritChoice_Product_Inventory.xlsx"]
    read_path = None
    for p in check_paths:
        if os.path.exists(p):
            read_path = p
            break
            
    if read_path:
        print(f"Reading existing stock, prices, and descriptions from '{read_path}' to preserve them...")
        try:
            old_wb = load_workbook(read_path, data_only=True)
            old_ws = old_wb.active
            for r_idx in range(5, old_ws.max_row + 1):
                sku = old_ws.cell(row=r_idx, column=1).value
                desc = old_ws.cell(row=r_idx, column=9).value
                stock = old_ws.cell(row=r_idx, column=10).value
                price = old_ws.cell(row=r_idx, column=11).value
                if sku:
                    existing_stock_price[sku] = (stock, price)
                    if desc:
                        existing_descriptions[sku] = desc
            print(f"Successfully loaded {len(existing_stock_price)} existing stock/price mappings and {len(existing_descriptions)} descriptions.")
        except Exception as e:
            print(f"Warning: Could not read existing stock/price/description data ({str(e)}). Preserving might not work.")
            
    print("Scanning products...")
    
    # List of all categories (directories inside Product_Images_1)
    categories = [d for d in os.listdir(image_dir) if os.path.isdir(os.path.join(image_dir, d))]
    
    inventory_data = []
    
    # Process each category
    for cat in sorted(categories):
        cat_path = os.path.join(image_dir, cat)
        files = [f for f in os.listdir(cat_path) if os.path.isfile(os.path.join(cat_path, f))]
        
        sku_prefix = SKU_PREFIXES.get(cat, "BC-GEN-")
        idx = 1
        
        for f in sorted(files):
            # Skip non-product files like resume PDF or thumbs.db
            if f.endswith(".pdf") or f.lower() == "thumbs.db" or not f.lower().endswith(('.jpeg', '.jpg', '.png', '.webp')):
                print(f"Skipping non-product file: {cat}/{f}")
                continue
                
            brand, prod_name, variant, qty, unit = parse_filename(f, cat)
            
            sku = f"{sku_prefix}{idx:03d}"
            idx += 1
            
            # Combine to make full product title
            size_phrase = f" ({qty} {unit})" if qty and unit else ""
            variant_phrase = f" - {variant}" if variant and variant.lower() != "original" else ""
            full_title = f"{brand} {prod_name}{variant_phrase}{size_phrase}"
            
            # Clean capitalization on title
            full_title = clean_capitalization(full_title)
            
            # Generate description (preserve existing if present, otherwise generate)
            description = existing_descriptions.get(sku)
            if not description:
                description = generate_product_description(brand, prod_name, variant, cat, qty, unit)
            
            # Relative image path
            img_path = f"{image_dir}/{cat}/{f}"
            
            inventory_data.append({
                "SKU": sku,
                "Category": clean_capitalization(cat.replace('_', ' ')),
                "Brand": brand,
                "Product Name": prod_name,
                "Variant": variant,
                "Size": qty,
                "Unit": unit,
                "Full Title": full_title,
                "Description": description,
                "Image Path": img_path
            })
            
    print(f"Successfully parsed {len(inventory_data)} products.")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Products Inventory"
    
    # Enable Gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    font_family = "Segoe UI"
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate `#1E293B`
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    
    data_font = Font(name=font_family, size=10)
    sku_font = Font(name=font_family, size=10, bold=True, color="1E3A8A") # Navy Blue for SKU
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Title Block (Rows 1 and 2)
    ws.merge_cells("A1:L2")
    title_cell = ws["A1"]
    title_cell.value = "BritChoice Kenya - UK Imports Inventory Database"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    
    # Row 3 is a spacer row, set height to 12
    ws.row_dimensions[3].height = 12
    for col_idx in range(1, 13):
        ws.cell(row=3, column=col_idx).fill = white_fill
        
    # Headers Row (Row 4)
    headers = [
        "Product SKU", "Category", "Brand", "Product Name", "Variant", 
        "Size / Qty", "Unit", "Full Product Title", "Description", 
        "Stock Count", "Price (KES)", "Image File Path"
    ]
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 6, 7, 10, 11] else "left", vertical="center")
    
    ws.row_dimensions[4].height = 28
    
    # Insert product data (Starting Row 5)
    current_row = 5
    for item in inventory_data:
        # Convert Size to float/int if possible
        size_val = item["Size"]
        if size_val:
            try:
                size_val = float(size_val) if '.' in size_val else int(size_val)
            except ValueError:
                pass
                
        # Look up existing stock and price values
        stock_val = None
        price_val = None
        if item["SKU"] in existing_stock_price:
            stock_val, price_val = existing_stock_price[item["SKU"]]
            
        # Row values mapping
        row_vals = [
            item["SKU"],
            item["Category"],
            item["Brand"],
            item["Product Name"],
            item["Variant"],
            size_val,
            item["Unit"],
            item["Full Title"],
            item["Description"],
            stock_val,  # Restored Stock Count
            price_val,  # Restored Price (KES)
            item["Image Path"]
        ]
        
        is_zebra = (current_row % 2 == 0)
        row_fill = zebra_fill if is_zebra else white_fill
        
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = sku_font if col_idx == 1 else data_font
            cell.fill = row_fill
            cell.border = thin_border
            
            # Alignments & formatting (vertical="top" + wrap_text for cleaner layout)
            if col_idx == 1: # SKU
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx in [2, 6, 7]: # Category, Size, Unit
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx in [8, 9, 12]: # Full Title, Description, Image Path (wrapped)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col_idx in [10, 11]: # Stock, Price
                cell.alignment = Alignment(horizontal="right", vertical="top")
                if col_idx == 11:
                    cell.number_format = '[$KES] #,##0'
                else:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")
                
        ws.row_dimensions[current_row].height = 42 # Increased row height to fit wrapped descriptions elegantly
        current_row += 1
        
    # Explicit column widths to ensure a generous and spacious layout
    ws.column_dimensions['A'].width = 18  # SKU
    ws.column_dimensions['B'].width = 18  # Category
    ws.column_dimensions['C'].width = 22  # Brand
    ws.column_dimensions['D'].width = 35  # Product Name
    ws.column_dimensions['E'].width = 25  # Variant
    ws.column_dimensions['F'].width = 14  # Size / Qty
    ws.column_dimensions['G'].width = 14  # Unit
    ws.column_dimensions['H'].width = 50  # Full Product Title (Wrapped)
    ws.column_dimensions['I'].width = 85  # Description (Wrapped)
    ws.column_dimensions['J'].width = 15  # Stock Count
    ws.column_dimensions['K'].width = 18  # Price (KES)
    ws.column_dimensions['L'].width = 55  # Image File Path (Wrapped)
    
    try:
        wb.save(excel_filename)
        print(f"Excel file successfully generated and saved to '{excel_filename}'.")
    except PermissionError:
        print(f"\n[ERROR] Permission denied: '{excel_filename}'.")
        print("Please CLOSE the Excel file if you have it open, then run the script again.")
        import sys
        sys.exit(1)

if __name__ == "__main__":
    main()
